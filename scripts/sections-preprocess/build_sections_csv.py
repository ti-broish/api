#!/usr/bin/env python3
"""
Build sections.csv from CIK xlsx files for NS 19.04.2026.

Inputs (hardcoded paths — adjust if you rerun):
  - Domestic voters:    ~/Downloads/obsht_broi_sekcii_izbirateli_izbiratelni_predsrochni-НС-06-04-2026.xlsx
  - Domestic addresses: ~/Downloads/Adresi-Izb-Sekcii(2).xlsx
  - Abroad addresses:   ~/Downloads/Таблица с номерата на избирателните секции и адреси.xlsx
  - Abroad machines:    ~/Downloads/spisak mashini za glasuvane izvan stranata.xlsx

Output:
  src/seeds/sections.csv with columns matching the existing seed format.
"""
import csv
import os
import re
from pathlib import Path

import openpyxl

HOME = Path.home()
DOWNLOADS = HOME / 'Downloads'

VOTERS_XLSX = DOWNLOADS / 'obsht_broi_sekcii_izbirateli_izbiratelni_predsrochni-НС-06-04-2026.xlsx'
ADDRESSES_XLSX = DOWNLOADS / 'Adresi-Izb-Sekcii(2).xlsx'
ABROAD_XLSX = DOWNLOADS / 'Таблица с номерата на избирателните секции и адреси.xlsx'
ABROAD_MACHINES_XLSX = DOWNLOADS / 'spisak mashini za glasuvane izvan stranata (1).xlsx'
DOMESTIC_MACHINES_XLSX = DOWNLOADS / 'СИК над 299 избиратели за 19.04.2026.xlsx'

# Last election open data — used to carry over mobile/ship sections that CIK
# typically adds after the initial registers are published.
LAST_ELECTION_SECTIONS = HOME / 'Downloads' / 'Актуализирана база данни' / 'sections.txt'

OUT_CSV = Path(__file__).resolve().parent.parent.parent / 'src' / 'seeds' / 'sections.csv'

CSV_COLUMNS = [
    'section_id', 'country_code', 'country_name',
    'election_region_name', 'election_region_code',
    'municipality_code', 'municipality_name',
    'city_region_code', 'city_region_name',
    'section_code', 'town_code', 'town_name', 'place',
    'voters_count', 'machines_count',
    'is_mobile', 'is_ship', 'is_covid',
]

# Canonical ER names (from last cycle's prep spreadsheet). CIK voters file uses
# adjective forms like "БЛАГОЕВГРАДСКИ"; we override to the noun form used in
# our historical seed data. Sofia oblast maps to three ERs distinguished by
# district; ER 16/17 distinguish Пловдив град vs област.
CANONICAL_ER_NAMES = {
    '01': 'Благоевград',
    '02': 'Бургас',
    '03': 'Варна',
    '04': 'Велико Търново',
    '05': 'Видин',
    '06': 'Враца',
    '07': 'Габрово',
    '08': 'Добрич',
    '09': 'Кърджали',
    '10': 'Кюстендил',
    '11': 'Ловеч',
    '12': 'Монтана',
    '13': 'Пазарджик',
    '14': 'Перник',
    '15': 'Плевен',
    '16': 'Пловдив град',
    '17': 'Пловдив област',
    '18': 'Разград',
    '19': 'Русе',
    '20': 'Силистра',
    '21': 'Сливен',
    '22': 'Смолян',
    '23': 'София 23 МИР',
    '24': 'София 24 МИР',
    '25': 'София 25 МИР',
    '26': 'София област',
    '27': 'Стара Загора',
    '28': 'Търговище',
    '29': 'Хасково',
    '30': 'Шумен',
    '31': 'Ямбол',
    '32': 'Извън страната',
}


def titlecase_bg(s: str) -> str:
    """CIK data is ALL CAPS. Convert to Title Case, keep single-letter words as-is."""
    if not s:
        return s
    return ' '.join(w.capitalize() if len(w) > 1 else w for w in s.strip().split())


def normalize_town_name(raw: str) -> str:
    """
    Strip ГР./С./К.С./etc. prefix, title-case the rest, prepend 'гр. ' or 'с. ' etc.
    e.g. 'ГР.БАНСКО' -> 'гр. Банско'
         'С.ДОБРИНИЩЕ' -> 'с. Добринище'
    """
    if not raw:
        return raw
    raw = raw.strip()
    m = re.match(r'^(ГР\.|С\.|К\.С\.|МАНАСТИР|СЕЛО|ГРАД)\s*(.+)$', raw, re.IGNORECASE)
    if m:
        prefix, name = m.group(1).upper(), m.group(2).strip()
        prefix_map = {
            'ГР.': 'гр.',
            'С.': 'с.',
            'К.С.': 'к.с.',
            'МАНАСТИР': 'манастир',
            'СЕЛО': 'с.',
            'ГРАД': 'гр.',
        }
        short = prefix_map.get(prefix, prefix.lower())
        return f'{short} {titlecase_bg(name)}'
    return titlecase_bg(raw)


def clean(v):
    if v is None:
        return ''
    return str(v).strip()


def parse_voters():
    """
    Return (voters_by_key, er_info, er_by_location):
      - voters_by_key: (oblast, muni, district, sec) -> dict with voters_count, names, er_code
      - er_info: er_code -> er_name
      - er_by_location: (oblast, muni, district) -> er_code
    """
    wb = openpyxl.load_workbook(VOTERS_XLSX, read_only=True, data_only=True)
    ws = wb.active
    voters = {}
    er_info = {}
    er_by_location = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:  # skip title + header
            continue
        flag, er_desc, oblast, oblast_name, muni, muni_name, district, district_name, sec_code, voters_count, *_ = row
        oblast = clean(oblast)
        if not oblast or not clean(sec_code):
            continue
        # ER code + name: col 1 is "XX - NAME"
        m = re.match(r'^(\d+)\s*-\s*(.+)$', clean(er_desc))
        if not m:
            continue
        er_code = m.group(1).zfill(2)
        er_name = titlecase_bg(m.group(2))
        er_info[er_code] = er_name
        muni_code = clean(muni) or '00'
        district_code = clean(district) or '00'
        sec_code = clean(sec_code).zfill(3)
        er_by_location[(oblast, muni_code, district_code)] = er_code
        key = (oblast, muni_code, district_code, sec_code)
        voters[key] = {
            'voters_count': voters_count,
            'oblast_name': titlecase_bg(clean(oblast_name)),
            'muni_name': titlecase_bg(clean(muni_name)),
            'district_name': titlecase_bg(clean(district_name)) if district_name else '',
            'er_code': er_code,
            'er_name': er_name,
        }
    return voters, er_info, er_by_location


def parse_addresses():
    """Return dict keyed by (oblast, municipality, district, section) -> address info."""
    wb = openpyxl.load_workbook(ADDRESSES_XLSX, read_only=True, data_only=True)
    ws = wb.active
    addrs = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # header
            continue
        oblast, oblast_name, muni, muni_name, district, district_name, sec_code, town_code, town_name, address = row
        oblast = clean(oblast)
        if not oblast or not clean(sec_code):
            continue
        muni_code = clean(muni) or '00'
        district_code = clean(district) or '00'
        sec_code = clean(sec_code).zfill(3)
        key = (oblast, muni_code, district_code, sec_code)
        addrs[key] = {
            'town_code': clean(town_code),
            'town_name': normalize_town_name(clean(town_name)),
            'address': clean(address),
            'district_name': titlecase_bg(clean(district_name)) if district_name else '',
            'muni_name': titlecase_bg(clean(muni_name)),
            'oblast_name': titlecase_bg(clean(oblast_name)),
        }
    return addrs


def parse_domestic_machines():
    """
    Return set of domestic section_ids that have a voting machine.
    The rule is >299 voters → 1 machine. File lists all qualifying sections.
    """
    wb = openpyxl.load_workbook(DOMESTIC_MACHINES_XLSX, read_only=True, data_only=True)
    ws = wb.active
    sids = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        sid = clean(row[0])
        if sid and len(sid) == 9 and not sid.startswith('32'):
            sids.add(sid)
    return sids


def parse_abroad_machines():
    """Return set of section_ids that have machines."""
    wb = openpyxl.load_workbook(ABROAD_MACHINES_XLSX, read_only=True, data_only=True)
    ws = wb.active
    machines = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # header
            continue
        country, place, count, sec_id = row
        sec_id = clean(sec_id)
        if not sec_id or not sec_id.startswith('32'):
            continue
        try:
            machines[sec_id] = int(count) if count else 1
        except (ValueError, TypeError):
            machines[sec_id] = 1
    return machines


def parse_abroad():
    """
    Parse the abroad addresses file. The structure is:
      row with country set, place, count, section_id, address_bg, ...
      next rows: country=None (continuation), place, count, section_id, ...
      summary rows like 'Общо: X; в ДКП: Y; извън ДКП: Z.' should be skipped.
    Returns list of dicts with: section_id, country_name, place, address.
    """
    wb = openpyxl.load_workbook(ABROAD_XLSX, read_only=True, data_only=True)
    ws = wb.active
    sections = []
    current_country = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        country, place, count, sec_id, addr_bg, addr_local, *_ = row
        country = clean(country)
        place = clean(place)
        sec_id = clean(sec_id)
        addr_bg = clean(addr_bg)

        # Skip empty / summary rows
        if country and country.startswith('Общо:'):
            continue
        if not sec_id or not sec_id.startswith('32'):
            # But update current_country if first-column looks like a country name
            if country and not country.startswith('Общо:') and not country.startswith('Таблица'):
                current_country = country
            continue

        if country:
            current_country = country

        # Strip CIK annotation asterisk and any trailing "(…)" parenthetical
        country_name = (current_country or '').rstrip(' *').rstrip()
        country_name = re.sub(r'\s*\([^)]*\)\s*$', '', country_name).strip()

        sections.append({
            'section_id': sec_id,
            'country_name': country_name,
            'place': place,
            'address': addr_bg,
        })
    return sections


def parse_last_election_mobile_ship():
    """
    Return list of (section_id, town_code, town_name, place, is_mobile, is_ship, machines)
    for mobile or ship sections from last election. These get carried over because
    CIK doesn't publish updated registers after late additions.
    """
    if not LAST_ELECTION_SECTIONS.exists():
        print(f'WARN: last election data not found at {LAST_ELECTION_SECTIONS}')
        return []
    out = []
    with open(LAST_ELECTION_SECTIONS, encoding='utf-8') as f:
        for line in f:
            fields = line.rstrip('\n').split(';')
            if len(fields) < 10:
                continue
            sid = fields[0].strip()
            is_mobile = fields[6].strip() == '1'
            is_ship = fields[7].strip() == '1'
            if not (is_mobile or is_ship):
                continue
            out.append({
                'section_id': sid,
                'town_code': fields[3].strip(),
                'town_name': fields[4].strip(),
                'place': fields[5].strip(),
                'is_mobile': is_mobile,
                'is_ship': is_ship,
                'machines_count': fields[8].strip() or '0',
            })
    return out


def build_rows():
    voters, er_info, er_by_location = parse_voters()
    addrs = parse_addresses()
    machines = parse_abroad_machines()
    domestic_machines = parse_domestic_machines()
    abroad = parse_abroad()
    last_mobile = parse_last_election_mobile_ship()

    rows = []

    # --- Domestic ---
    # Union of keys from both files
    keys = set(voters.keys()) | set(addrs.keys())
    missing_voters = []
    missing_addrs = []
    missing_er = []
    for key in sorted(keys):
        oblast, muni, district, sec = key
        v = voters.get(key, {})
        a = addrs.get(key, {})
        if not v:
            missing_voters.append(key)
        if not a:
            missing_addrs.append(key)

        er_code = v.get('er_code') or er_by_location.get((oblast, muni, district))
        if not er_code:
            missing_er.append(key)
            continue
        er_name = CANONICAL_ER_NAMES.get(er_code) or er_info.get(er_code) or v.get('er_name') or titlecase_bg(a.get('oblast_name', ''))
        muni_name = v.get('muni_name') or a.get('muni_name') or ''
        district_name = v.get('district_name') or a.get('district_name') or ''
        town_code = a.get('town_code', '')
        town_name = a.get('town_name', '')
        place = a.get('address', '')
        voters_count = v.get('voters_count') or ''

        section_id = f'{er_code}{muni}{district}{sec}'  # 2+2+2+3 = 9 chars

        rows.append({
            'section_id': section_id,
            'country_code': '000',
            'country_name': 'България',
            'election_region_name': er_name,
            'election_region_code': er_code,
            'municipality_code': muni,
            'municipality_name': muni_name,
            'city_region_code': district,
            'city_region_name': district_name,
            'section_code': sec,
            'town_code': town_code,
            'town_name': town_name,
            'place': place,
            'voters_count': voters_count,
            'machines_count': '1' if section_id in domestic_machines else '0',
            'is_mobile': '',
            'is_ship': '',
            'is_covid': '',
        })

    # --- Abroad ---
    # Assign sequential town codes starting at 100000 per unique place name
    abroad_town_code_map = {}
    next_town_code = 100000
    for s in sorted(abroad, key=lambda x: x['section_id']):
        sec_id = s['section_id']  # 9 chars: 32 + country(3) + 0 + seq(3)
        er_code = sec_id[:2]
        country_code = sec_id[2:5]
        sec_code = sec_id[6:9]

        # Split "City (Tag) [trailing]" → city in town_name, tag + trailing merged into place.
        # Handles: "Лондон (Посолство 1)", "Анкара (Посолство) 1", "Баку (Посолство) *",
        # "Виена (Посолски комплекс) - ПК Нашмаркт".
        raw_place = s['place'].rstrip(' *').rstrip()  # drop CIK annotation asterisk
        m = re.search(r'\s*\(([^)]+)\)\s*', raw_place)
        if m:
            city = raw_place[:m.start()].strip()
            trailing = raw_place[m.end():].strip().lstrip('-').strip()
            tag = f'{m.group(1).strip()} {trailing}'.strip()
        else:
            city, tag = raw_place, ''

        addr = s['address']
        if tag and addr:
            place_out = f'{tag}, {addr}'
        else:
            place_out = tag or addr

        # Town code keyed by (country, city) so all "Лондон" sections share one town row
        town_key = (country_code, city)
        if town_key not in abroad_town_code_map:
            abroad_town_code_map[town_key] = next_town_code
            next_town_code += 1
        town_code = abroad_town_code_map[town_key]

        rows.append({
            'section_id': sec_id,
            'country_code': country_code,
            'country_name': s['country_name'],
            'election_region_name': 'Извън страната',
            'election_region_code': er_code,
            'municipality_code': '00',
            'municipality_name': '',
            'city_region_code': '00',
            'city_region_name': '',
            'section_code': sec_code,
            'town_code': str(town_code),
            'town_name': city,
            'place': place_out,
            'voters_count': '',
            'machines_count': str(machines.get(sec_id, 0)),
            'is_mobile': '',
            'is_ship': '',
            'is_covid': '',
        })

    # --- Carry-over mobile/ship from last election ---
    # Build lookups from domestic rows for ER/municipality/city_region names
    existing_ids = {r['section_id'] for r in rows}
    er_name_by_code = {}
    muni_name_by_code = {}
    cr_name_by_code = {}
    for r in rows:
        if r['election_region_code'] == '32':
            continue
        er_name_by_code[r['election_region_code']] = r['election_region_name']
        muni_name_by_code[(r['election_region_code'], r['municipality_code'])] = r['municipality_name']
        if r['city_region_code'] != '00':
            cr_name_by_code[(r['election_region_code'], r['municipality_code'], r['city_region_code'])] = r['city_region_name']

    added_mobile = 0
    added_ship = 0
    skipped = 0
    for s in last_mobile:
        if s['section_id'] in existing_ids:
            continue
        sid = s['section_id']
        er_code = sid[0:2]
        muni_code = sid[2:4]
        cr_code = sid[4:6]
        sec_code = sid[6:9]

        er_name = er_name_by_code.get(er_code)
        muni_name = muni_name_by_code.get((er_code, muni_code), '')
        cr_name = cr_name_by_code.get((er_code, muni_code, cr_code), '') if cr_code != '00' else ''

        if not er_name:
            skipped += 1
            continue

        rows.append({
            'section_id': sid,
            'country_code': '000',
            'country_name': 'България',
            'election_region_name': er_name,
            'election_region_code': er_code,
            'municipality_code': muni_code,
            'municipality_name': muni_name,
            'city_region_code': cr_code,
            'city_region_name': cr_name,
            'section_code': sec_code,
            'town_code': s['town_code'],
            'town_name': s['town_name'],
            'place': s['place'],
            'voters_count': '',
            'machines_count': s['machines_count'],
            'is_mobile': 't' if s['is_mobile'] else '',
            'is_ship': 't' if s['is_ship'] else '',
            'is_covid': '',
        })
        if s['is_mobile']:
            added_mobile += 1
        if s['is_ship']:
            added_ship += 1

    print(f'Domestic sections: {sum(1 for r in rows if r["election_region_code"] != "32")}')
    print(f'Abroad sections:   {sum(1 for r in rows if r["election_region_code"] == "32")}')
    print(f'  carried-over mobile: {added_mobile}')
    print(f'  carried-over ship:   {added_ship}')
    print(f'  skipped (no ER match): {skipped}')
    print(f'Missing voters:    {len(missing_voters)}')
    print(f'Missing addresses: {len(missing_addrs)}')
    if missing_voters[:5]:
        print(f'  sample missing voters: {missing_voters[:5]}')
    if missing_addrs[:5]:
        print(f'  sample missing addrs: {missing_addrs[:5]}')

    return rows


def main():
    rows = build_rows()
    # csvToSql in the seed does a naive .split(';') on its output, so any
    # embedded semicolon in a string value breaks the SQL. Replace with commas.
    for r in rows:
        for k in ('place', 'town_name', 'municipality_name', 'city_region_name', 'country_name'):
            if r.get(k) and ';' in r[k]:
                r[k] = r[k].replace(';', ',')
    with open(OUT_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS, quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()
        for r in rows:
            writer.writerow(r)
    print(f'\nWrote {len(rows)} rows to {OUT_CSV}')


if __name__ == '__main__':
    main()
